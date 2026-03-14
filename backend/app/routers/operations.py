"""
Operations Intelligence API - admin-only. Usage analytics, heatmaps, YUCG resources, Ollama insights, Excel export.
Private and internal only.
"""
import io
import json
import os
from datetime import datetime, timedelta
from fastapi import APIRouter, Depends, HTTPException, Query, UploadFile, File
from fastapi.responses import Response
from pydantic import BaseModel
from typing import Optional

from app.database import get_db, row_to_dict
from app.auth_deps import get_current_admin
from app.services.usage_service import log_event

router = APIRouter()

# YUCG Excel palette
YUCG_HEADER = "1A2F5A"
YUCG_BORDER = "C8DCED"
YUCG_ALT = "E8EEF4"


# --- Usage events (admin only) ---
@router.get("/events")
async def list_events(
    admin: dict = Depends(get_current_admin),
    limit: int = Query(500, le=2000),
    event_type: Optional[str] = None,
    days: int = Query(30, le=365),
):
    """List usage events for analytics. Admin only."""
    db = await get_db()
    try:
        since = (datetime.utcnow() - timedelta(days=days)).isoformat().replace("T", " ")
        if event_type:
            cursor = await db.execute(
                """SELECT id, user_id, event_type, resource_type, details_json, created_at
                   FROM usage_events WHERE created_at >= ? AND event_type = ?
                   ORDER BY created_at DESC LIMIT ?""",
                (since, event_type, limit),
            )
        else:
            cursor = await db.execute(
                """SELECT id, user_id, event_type, resource_type, details_json, created_at
                   FROM usage_events WHERE created_at >= ?
                   ORDER BY created_at DESC LIMIT ?""",
                (since, limit),
            )
        rows = await cursor.fetchall()
        out = []
        for r in rows:
            d = dict(r)
            if d.get("details_json"):
                try:
                    d["details"] = json.loads(d["details_json"])
                except Exception:
                    d["details"] = None
            del d["details_json"]
            out.append(d)
        return out
    finally:
        await db.close()


@router.get("/heatmap")
async def get_heatmap(
    admin: dict = Depends(get_current_admin),
    days: int = Query(30, le=365),
    group_by: str = Query("hour", description="hour or day_of_week"),
):
    """Aggregate events for heatmap: counts by hour (0-23) or day of week (0-6). Admin only.
    Returns grid (bucket -> event_type -> count) plus matrix_2d for horizontal display: rows = time buckets, cols = event types."""
    db = await get_db()
    try:
        since = (datetime.utcnow() - timedelta(days=days)).isoformat().replace("T", " ")
        if group_by == "day_of_week":
            cursor = await db.execute(
                """SELECT CAST(strftime('%w', created_at) AS INT) AS bucket, event_type, COUNT(*) AS count
                   FROM usage_events WHERE created_at >= ?
                   GROUP BY bucket, event_type ORDER BY bucket, event_type""",
                (since,),
            )
        else:
            cursor = await db.execute(
                """SELECT CAST(strftime('%H', created_at) AS INT) AS bucket, event_type, COUNT(*) AS count
                   FROM usage_events WHERE created_at >= ?
                   GROUP BY bucket, event_type ORDER BY bucket, event_type""",
                (since,),
            )
        rows = await cursor.fetchall()
        grid = {}
        for r in rows:
            b = r["bucket"]
            if b not in grid:
                grid[b] = {}
            grid[b][r["event_type"]] = r["count"]
        # Build 2D matrix for horizontal heatmap: row_labels (hours or days), col_labels (event types), values[row][col]
        buckets = sorted(grid.keys())
        all_types = set()
        for g in grid.values():
            all_types.update(g.keys())
        col_labels = sorted(all_types)
        row_labels = [str(b) for b in buckets]
        values = []
        for b in buckets:
            row = [grid[b].get(et, 0) for et in col_labels]
            values.append(row)
        return {
            "group_by": group_by,
            "days": days,
            "grid": grid,
            "rows": [dict(r) for r in rows],
            "matrix_2d": {"row_labels": row_labels, "col_labels": col_labels, "values": values},
        }
    finally:
        await db.close()


@router.get("/cursor-heatmap")
async def get_cursor_heatmap(
    admin: dict = Depends(get_current_admin),
    days: int = Query(30, le=365),
    bins: int = Query(10, ge=5, le=20),
):
    """Aggregate cursor events into 2D grids per page (where users' cursors go). Admin only."""
    db = await get_db()
    try:
        since = (datetime.utcnow() - timedelta(days=days)).isoformat().replace("T", " ")
        cursor = await db.execute(
            """SELECT resource_type, details_json FROM usage_events
               WHERE event_type = 'cursor' AND created_at >= ? AND details_json IS NOT NULL""",
            (since,),
        )
        rows = await cursor.fetchall()
    finally:
        await db.close()

    # Per-page: (x_bin, y_bin) -> count. x,y are expected 0-100 in details.
    pages = {}
    for r in rows:
        try:
            d = json.loads(r["details_json"] or "{}")
            x = d.get("x")
            y = d.get("y")
            if x is None or y is None:
                continue
            x, y = float(x), float(y)
            page = (r["resource_type"] or "").strip() or "unknown"
            if page not in pages:
                pages[page] = {}
            bx = min(int(x * bins / 100), bins - 1) if 0 <= x <= 100 else 0
            by = min(int(y * bins / 100), bins - 1) if 0 <= y <= 100 else 0
            key = (bx, by)
            pages[page][key] = pages[page].get(key, 0) + 1
        except (ValueError, TypeError):
            continue

    # Convert to 2D array per page
    result = {}
    for page, counts in pages.items():
        grid = [[0] * bins for _ in range(bins)]
        for (bx, by), c in counts.items():
            if 0 <= by < bins and 0 <= bx < bins:
                grid[by][bx] = c
        result[page] = {"grid": grid, "bins": bins}
    return {"days": days, "pages": result}


@router.get("/aggregates")
async def get_aggregates(
    admin: dict = Depends(get_current_admin),
    days: int = Query(30, le=365),
):
    """Event counts by type and by resource_type for charts. Admin only."""
    db = await get_db()
    try:
        since = (datetime.utcnow() - timedelta(days=days)).isoformat().replace("T", " ")
        cursor = await db.execute(
            """SELECT event_type, COUNT(*) AS count FROM usage_events WHERE created_at >= ?
               GROUP BY event_type ORDER BY count DESC""",
            (since,),
        )
        by_type = [dict(r) for r in await cursor.fetchall()]
        cursor = await db.execute(
            """SELECT resource_type, COUNT(*) AS count FROM usage_events WHERE created_at >= ? AND resource_type != ''
               GROUP BY resource_type ORDER BY count DESC""",
            (since,),
        )
        by_resource = [dict(r) for r in await cursor.fetchall()]
        return {"by_event_type": by_type, "by_resource_type": by_resource, "days": days}
    finally:
        await db.close()


# --- YUCG resources (upload for Ollama context) ---
@router.get("/resources")
async def list_resources(admin: dict = Depends(get_current_admin)):
    """List ingested YUCG resources. Admin only."""
    db = await get_db()
    try:
        cursor = await db.execute(
            "SELECT id, name, content_type, LENGTH(content_text) AS content_length, created_at FROM yucg_resources ORDER BY created_at DESC"
        )
        rows = await cursor.fetchall()
        return [dict(r) for r in rows]
    finally:
        await db.close()


class ResourceCreate(BaseModel):
    name: str
    content_text: str
    content_type: Optional[str] = "text"


@router.post("/resources")
async def create_resource(payload: ResourceCreate, admin: dict = Depends(get_current_admin)):
    """Add a YUCG resource (e.g. pasted text or doc content). Admin only."""
    db = await get_db()
    try:
        await db.execute(
            "INSERT INTO yucg_resources (name, content_type, content_text) VALUES (?, ?, ?)",
            (payload.name, payload.content_type or "text", payload.content_text),
        )
        await db.commit()
        return {"ok": True}
    finally:
        await db.close()


@router.post("/resources/upload")
async def upload_resource(file: UploadFile = File(...), admin: dict = Depends(get_current_admin)):
    """Upload a text/PDF file as a YUCG resource. Content stored as text. Admin only."""
    content = await file.read()
    name = file.filename or "uploaded"
    # Try decode as text; if binary, store as base64 or skip
    try:
        text = content.decode("utf-8", errors="replace")
    except Exception:
        text = f"[Binary file: {name}, size {len(content)} bytes]"
    db = await get_db()
    try:
        await db.execute(
            "INSERT INTO yucg_resources (name, content_type, content_text) VALUES (?, ?, ?)",
            (name, "upload", text[:500_000]),  # cap size
        )
        await db.commit()
        return {"ok": True}
    finally:
        await db.close()


# --- Ollama operations query (internal learning from resources + usage) ---
class OllamaQuery(BaseModel):
    query: str


@router.post("/ollama/query")
async def ollama_operations_query(payload: OllamaQuery, admin: dict = Depends(get_current_admin)):
    """Ask Ollama about YUCG operations using ingested resources + recent usage summary. Admin only. Runs internally."""
    db = await get_db()
    try:
        cursor = await db.execute(
            "SELECT name, content_text FROM yucg_resources ORDER BY created_at DESC LIMIT 20"
        )
        resources = await cursor.fetchall()
        resource_context = "\n\n".join(
            f"--- {r['name']} ---\n{(r['content_text'] or '')[:8000]}" for r in resources
        )[:30000]

        cursor = await db.execute(
            """SELECT event_type, resource_type, details_json, created_at FROM usage_events
               ORDER BY created_at DESC LIMIT 500"""
        )
        events = await cursor.fetchall()
        usage_summary = []
        for e in events[:200]:
            usage_summary.append(
                f"{e['created_at']} | {e['event_type']} | {e['resource_type']} | {e['details_json'] or ''}"
            )
        usage_context = "\n".join(usage_summary)[:15000]

        prompt = f"""You are an internal operations analyst for YUCG (Yale Undergraduate Consulting Group). Use ONLY the following context to answer. Be concise and data-driven.

YUCG RESOURCES (uploaded docs):
{resource_context or '(No resources uploaded yet)'}

RECENT USAGE EVENTS (website user actions):
{usage_context or '(No events yet)'}

QUESTION: {payload.query}

Answer based only on the above. If the data does not contain enough information, say so. Focus on patterns, efficiency, and what leadership could do to streamline operations."""

        ollama_url = os.getenv("OLLAMA_URL", "http://localhost:11434")
        import httpx
        async with httpx.AsyncClient(timeout=120.0) as client:
            r = await client.post(
                f"{ollama_url.rstrip('/')}/api/generate",
                json={"model": os.getenv("OLLAMA_MODEL", "llama3.2"), "prompt": prompt, "stream": False},
            )
        if r.status_code != 200:
            return {"answer": None, "error": f"Ollama returned {r.status_code}", "context_used": True}
        data = r.json()
        answer = (data.get("response") or "").strip()
        return {"answer": answer, "error": None, "context_used": True}
    finally:
        await db.close()


# --- Excel export: insights + user behavior patterns ---
@router.get("/export/insights", response_class=Response)
async def export_insights_excel(
    admin: dict = Depends(get_current_admin),
    days: int = Query(30, le=365),
):
    """Export usage aggregates and behavior patterns to a styled Excel file. Admin only."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    db = await get_db()
    try:
        since = (datetime.utcnow() - timedelta(days=days)).isoformat().replace("T", " ")
        cursor = await db.execute(
            """SELECT event_type, resource_type, COUNT(*) AS count FROM usage_events
               WHERE created_at >= ? GROUP BY event_type, resource_type ORDER BY event_type, count DESC""",
            (since,),
        )
        agg = await cursor.fetchall()
        cursor = await db.execute(
            """SELECT user_id, event_type, resource_type, created_at FROM usage_events
               WHERE created_at >= ? ORDER BY created_at DESC LIMIT 3000""",
            (since,),
        )
        events = await cursor.fetchall()
    finally:
        await db.close()

    wb = Workbook()
    # Sheet 1: Aggregates
    ws1 = wb.active
    ws1.title = "Behavior patterns"
    headers1 = ["Event type", "Resource type", "Count"]
    for col, h in enumerate(headers1, 1):
        c = ws1.cell(row=1, column=col, value=h)
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = PatternFill(start_color=YUCG_HEADER, end_color=YUCG_HEADER, fill_type="solid")
    for row_idx, r in enumerate(agg, 2):
        ws1.cell(row=row_idx, column=1, value=r["event_type"])
        ws1.cell(row=row_idx, column=2, value=r["resource_type"] or "")
        ws1.cell(row=row_idx, column=3, value=r["count"])
    for col in range(1, 4):
        ws1.column_dimensions[get_column_letter(col)].width = 22

    # Sheet 2: Raw events sample
    ws2 = wb.create_sheet("Raw events (sample)")
    headers2 = ["User ID", "Event type", "Resource type", "Created at"]
    for col, h in enumerate(headers2, 1):
        c = ws2.cell(row=1, column=col, value=h)
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = PatternFill(start_color=YUCG_HEADER, end_color=YUCG_HEADER, fill_type="solid")
    for row_idx, r in enumerate(events, 2):
        ws2.cell(row=row_idx, column=1, value=r["user_id"])
        ws2.cell(row=row_idx, column=2, value=r["event_type"])
        ws2.cell(row=row_idx, column=3, value=r["resource_type"] or "")
        ws2.cell(row=row_idx, column=4, value=str(r["created_at"]))
    for col in range(1, 5):
        ws2.column_dimensions[get_column_letter(col)].width = 20

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return Response(
        content=buf.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=YUCG_operations_insights_{days}d.xlsx"},
    )


# --- Matplotlib chart export (PNG) ---
@router.get("/export/charts", response_class=Response)
async def export_charts(
    admin: dict = Depends(get_current_admin),
    days: int = Query(30, le=365),
):
    """Export operations data as matplotlib PNG charts (event types, usage by hour). Admin only."""
    import zipfile

    db = await get_db()
    try:
        since = (datetime.utcnow() - timedelta(days=days)).isoformat().replace("T", " ")
        cursor = await db.execute(
            """SELECT event_type, COUNT(*) AS count FROM usage_events
               WHERE created_at >= ? GROUP BY event_type ORDER BY count DESC LIMIT 20""",
            (since,),
        )
        by_type_raw = await cursor.fetchall()
        cursor = await db.execute(
            """SELECT CAST(strftime('%H', created_at) AS INT) AS hour, COUNT(*) AS count
               FROM usage_events WHERE created_at >= ?
               GROUP BY hour ORDER BY hour""",
            (since,),
        )
        by_hour_raw = await cursor.fetchall()
    finally:
        await db.close()

    by_type = [dict(r) for r in by_type_raw]
    by_hour = [dict(r) for r in by_hour_raw]

    try:
        import matplotlib
        matplotlib.use("Agg")
        import matplotlib.pyplot as plt
    except Exception as e:
        raise HTTPException(status_code=503, detail=f"matplotlib not available: {e!s}")

    zip_buf = io.BytesIO()
    try:
        with zipfile.ZipFile(zip_buf, "w", zipfile.ZIP_DEFLATED) as zf:
            navy = "#1a2f5a"
            steel = "#5b7fa6"

            if by_type:
                fig, ax = plt.subplots(figsize=(10, 5))
                labels = [str(r.get("event_type") or "unknown")[:24] for r in by_type]
                counts = [int(r.get("count") or 0) for r in by_type]
                ax.bar(range(len(labels)), counts, color=steel, edgecolor=navy, linewidth=0.5)
                ax.set_xticks(range(len(labels)))
                ax.set_xticklabels(labels, rotation=45, ha="right")
                ax.set_ylabel("Count")
                ax.set_title(f"Usage events by type (last {days} days)")
                ax.set_facecolor("#f8fafc")
                fig.patch.set_facecolor("#f8fafc")
                plt.tight_layout()
                img_buf = io.BytesIO()
                fig.savefig(img_buf, format="png", dpi=120, bbox_inches="tight")
                plt.close(fig)
                img_buf.seek(0)
                zf.writestr("events_by_type.png", img_buf.getvalue())

            if by_hour:
                fig, ax = plt.subplots(figsize=(10, 4))
                hour_counts = [0] * 24
                for r in by_hour:
                    h = int(r["hour"]) if r.get("hour") is not None else 0
                    if 0 <= h < 24:
                        hour_counts[h] = int(r.get("count") or 0)
                ax.bar(range(24), hour_counts, color=steel, edgecolor=navy, linewidth=0.5)
                ax.set_xticks(range(0, 24, 2))
                ax.set_xlabel("Hour of day (UTC)")
                ax.set_ylabel("Count")
                ax.set_title(f"Usage by hour (last {days} days)")
                ax.set_facecolor("#f8fafc")
                fig.patch.set_facecolor("#f8fafc")
                plt.tight_layout()
                img_buf = io.BytesIO()
                fig.savefig(img_buf, format="png", dpi=120, bbox_inches="tight")
                plt.close(fig)
                img_buf.seek(0)
                zf.writestr("usage_by_hour.png", img_buf.getvalue())

            if not by_type and not by_hour:
                fig, ax = plt.subplots(figsize=(6, 3))
                ax.text(0.5, 0.5, "No usage data in this period", ha="center", va="center", fontsize=14)
                ax.set_facecolor("#f8fafc")
                fig.patch.set_facecolor("#f8fafc")
                img_buf = io.BytesIO()
                fig.savefig(img_buf, format="png", dpi=120)
                plt.close(fig)
                img_buf.seek(0)
                zf.writestr("no_data.png", img_buf.getvalue())
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Chart generation failed: {e!s}")

    zip_buf.seek(0)
    return Response(
        content=zip_buf.getvalue(),
        media_type="application/zip",
        headers={"Content-Disposition": f"attachment; filename=YUCG_operations_charts_{days}d.zip"},
    )


# --- Full export: Excel + charts + cursor cache + page snapshots (everything in one ZIP) ---
async def build_operations_full_export_bytes(admin: dict, days: int = 30) -> bytes:
    """Build the full operations export ZIP bytes. Used by export_full and admin export/all."""
    import zipfile

    db = await get_db()
    since = (datetime.utcnow() - timedelta(days=days)).isoformat().replace("T", " ")
    try:
        cursor = await db.execute(
            """SELECT event_type, resource_type, details_json, created_at FROM usage_events
               WHERE created_at >= ? ORDER BY created_at DESC LIMIT 5000""",
            (since,),
        )
        all_events_raw = await cursor.fetchall()
        cursor = await db.execute(
            """SELECT event_type, resource_type, COUNT(*) AS count FROM usage_events
               WHERE created_at >= ? GROUP BY event_type, resource_type""",
            (since,),
        )
        agg_raw = await cursor.fetchall()
        cursor = await db.execute(
            """SELECT CAST(strftime('%H', created_at) AS INT) AS hour, event_type, COUNT(*) AS count
               FROM usage_events WHERE created_at >= ? GROUP BY hour, event_type""",
            (since,),
        )
        by_hour_type_raw = await cursor.fetchall()
    finally:
        await db.close()

    all_events = [dict(r) for r in all_events_raw]
    agg = [dict(r) for r in agg_raw]
    by_hour_type = [dict(r) for r in by_hour_type_raw]

    try:
        import matplotlib
        matplotlib.use("Agg")
        import matplotlib.pyplot as plt
    except Exception as e:
        raise HTTPException(status_code=503, detail=f"matplotlib not available: {e!s}")

    zip_buf = io.BytesIO()
    navy, steel = "#1a2f5a", "#5b7fa6"

    try:
        with zipfile.ZipFile(zip_buf, "w", zipfile.ZIP_DEFLATED) as zf:
            # 1) Excel (same as insights)
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill
            from openpyxl.utils import get_column_letter
            wb = Workbook()
            ws1 = wb.active
            ws1.title = "Behavior patterns"
            for col, h in enumerate(["Event type", "Resource type", "Count"], 1):
                c = ws1.cell(row=1, column=col, value=h)
                c.font = Font(bold=True, color="FFFFFF")
                c.fill = PatternFill(start_color=YUCG_HEADER, end_color=YUCG_HEADER, fill_type="solid")
            for i, r in enumerate(agg, 2):
                ws1.cell(row=i, column=1, value=r.get("event_type") or "")
                ws1.cell(row=i, column=2, value=r.get("resource_type") or "")
                ws1.cell(row=i, column=3, value=int(r.get("count") or 0))
            for col in range(1, 4):
                ws1.column_dimensions[get_column_letter(col)].width = 22
            xlsx_buf = io.BytesIO()
            wb.save(xlsx_buf)
            xlsx_buf.seek(0)
            zf.writestr("YUCG_operations_insights.xlsx", xlsx_buf.getvalue())

            # 2) Page snapshots: how each page is used (event types and counts)
            page_snapshots = {}
            for r in agg:
                res = r.get("resource_type") or ""
                et = r.get("event_type") or "unknown"
                if res not in page_snapshots:
                    page_snapshots[res] = {}
                page_snapshots[res][et] = int(r.get("count") or 0)
            snapshots = {
                "days": days,
                "generated_at": datetime.utcnow().isoformat() + "Z",
                "pages": page_snapshots,
                "description": "Per-page counts of event types (page_view, cursor, campaign_created, etc.).",
            }
            zf.writestr("page_snapshots.json", json.dumps(snapshots, indent=2))

            # 3) Cursor heatmap cache: aggregated cursor positions per page
            cursor_grids = {}
            for r in all_events:
                if r.get("event_type") != "cursor" or not r.get("details_json"):
                    continue
                try:
                    d = json.loads(r["details_json"] or "{}")
                    x, y = d.get("x"), d.get("y")
                    if x is None or y is None:
                        continue
                    x, y = float(x), float(y)
                    page = (r.get("resource_type") or "").strip() or "unknown"
                    if page not in cursor_grids:
                        cursor_grids[page] = {}
                    bins = 10
                    bx = min(int(x * bins / 100), bins - 1) if 0 <= x <= 100 else 0
                    by = min(int(y * bins / 100), bins - 1) if 0 <= y <= 100 else 0
                    key = (bx, by)
                    cursor_grids[page][key] = cursor_grids[page].get(key, 0) + 1
                except (ValueError, TypeError):
                    continue
            cursor_export = {}
            for page, counts in cursor_grids.items():
                grid = [[0] * 10 for _ in range(10)]
                for (bx, by), c in counts.items():
                    if 0 <= by < 10 and 0 <= bx < 10:
                        grid[by][bx] = c
                cursor_export[page] = {"grid": grid, "bins": 10}
            zf.writestr("cursor_heatmap_cache.json", json.dumps({"days": days, "pages": cursor_export}, indent=2))

            # 4) Chart: events by type
            by_type = {}
            for r in agg:
                et = r.get("event_type") or "unknown"
                by_type[et] = by_type.get(et, 0) + int(r.get("count") or 0)
            if by_type:
                fig, ax = plt.subplots(figsize=(10, 5))
                labels = list(by_type.keys())[:20]
                counts = [by_type[k] for k in labels]
                ax.bar(range(len(labels)), counts, color=steel, edgecolor=navy, linewidth=0.5)
                ax.set_xticks(range(len(labels)))
                ax.set_xticklabels(labels, rotation=45, ha="right")
                ax.set_ylabel("Count")
                ax.set_title(f"Usage by event type (last {days} days)")
                ax.set_facecolor("#f8fafc")
                fig.patch.set_facecolor("#f8fafc")
                plt.tight_layout()
                img = io.BytesIO()
                fig.savefig(img, format="png", dpi=120, bbox_inches="tight")
                plt.close(fig)
                img.seek(0)
                zf.writestr("charts/events_by_type.png", img.getvalue())

            # 5) Chart: 2D heatmap (hour x event type) - horizontal
            hour_types = {}
            for r in by_hour_type:
                h = int(r["hour"]) if r.get("hour") is not None else 0
                if 0 <= h < 24:
                    if h not in hour_types:
                        hour_types[h] = {}
                    hour_types[h][r.get("event_type") or "unknown"] = int(r.get("count") or 0)
            if hour_types:
                all_etypes = sorted(set(et for row in hour_types.values() for et in row))
                if all_etypes:
                    data = []
                    for h in range(24):
                        row = [hour_types.get(h, {}).get(et, 0) for et in all_etypes]
                        data.append(row)
                    fig, ax = plt.subplots(figsize=(max(14, len(all_etypes) * 0.8), 8))
                    im = ax.imshow(data, aspect="auto", cmap="Blues", vmin=0)
                    ax.set_xticks(range(len(all_etypes)))
                    ax.set_xticklabels(all_etypes, rotation=45, ha="right")
                    ax.set_yticks(range(24))
                    ax.set_yticklabels([f"{h}:00" for h in range(24)])
                    ax.set_xlabel("Event type")
                    ax.set_ylabel("Hour (UTC)")
                    ax.set_title(f"Usage heatmap: hour × event type (last {days} days)")
                    plt.colorbar(im, ax=ax, label="Count")
                    plt.tight_layout()
                    img = io.BytesIO()
                    fig.savefig(img, format="png", dpi=120, bbox_inches="tight")
                    plt.close(fig)
                    img.seek(0)
                    zf.writestr("charts/heatmap_hour_by_type.png", img.getvalue())

            # 6) Cursor heatmap image per page (if we have cursor data)
            for page, data in cursor_export.items():
                grid = data["grid"]
                fig, ax = plt.subplots(figsize=(6, 5))
                ax.imshow(grid, aspect="auto", cmap="Blues", vmin=0)
                ax.set_title(f"Cursor positions: {page or 'unknown'} (last {days}d)")
                ax.set_xlabel("X (left → right)")
                ax.set_ylabel("Y (top → bottom)")
                plt.tight_layout()
                img = io.BytesIO()
                fig.savefig(img, format="png", dpi=100, bbox_inches="tight")
                plt.close(fig)
                img.seek(0)
                safe_name = (page or "unknown").replace("/", "_").replace("\\", "_")[:30]
                zf.writestr(f"cursor_heatmaps/{safe_name}.png", img.getvalue())

    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Full export failed: {e!s}")

    zip_buf.seek(0)
    return zip_buf.getvalue()


@router.get("/export/full", response_class=Response)
async def export_full(
    admin: dict = Depends(get_current_admin),
    days: int = Query(30, le=365),
):
    """Export full cache: Excel, matplotlib charts, cursor heatmap data, page snapshots. All in one ZIP. Admin only."""
    content = await build_operations_full_export_bytes(admin, days)
    return Response(
        content=content,
        media_type="application/zip",
        headers={"Content-Disposition": f"attachment; filename=YUCG_operations_full_export_{days}d.zip"},
    )
