"""
Contacts API - Contact Scraper & Discovery Engine
"""
import csv
import io
import os
import json
from fastapi import APIRouter, HTTPException, UploadFile, File, Depends
from app.database import get_db
from app.models import ContactCreate, ScrapeRequest, SearchPersonRequest
from app.auth_deps import get_current_user_optional
from app.services.audit_service import log_audit
from app.services.usage_service import log_event
from app.services.contact_scraper import (
    scrape_contacts_from_domain,
    extract_domain_from_company,
    infer_email_from_name,
    normalize_domain,
    sanitize_email,
)
from app.services.linkedin_scraper import scrape_linkedin_company

router = APIRouter()


def _normalize_name(name: str) -> str:
    """Normalize name for matching (lowercase, strip)."""
    return (name or "").lower().strip()


def _merge_contacts(
    domain_contacts: list[dict],
    linkedin_contacts: list[dict],
    company: str,
    domain: str,
    custom_patterns: list[str] | None = None,
) -> list[dict]:
    """Merge domain and LinkedIn contacts."""
    seen_emails = set()
    merged = []
    by_name = {}

    for c in domain_contacts:
        email = c.get("email")
        if email and email not in seen_emails:
            seen_emails.add(email)
            merged.append(dict(c))
            if c.get("name"):
                by_name[_normalize_name(c["name"])] = merged[-1]

    for li in linkedin_contacts:
        name = li.get("name")
        if not name:
            continue
        norm = _normalize_name(name)
        matched = by_name.get(norm)
        if matched:
            matched["linkedin_url"] = li.get("linkedin_url") or matched.get("linkedin_url")
            matched["title"] = matched.get("title") or li.get("title")
            continue
        email = li.get("email") or (infer_email_from_name(name, domain, custom_patterns) if domain else None)
        if not email:
            email = f"linkedin-{name.replace(' ', '.').lower()}@{domain or 'placeholder.local'}"
        if email in seen_emails:
            continue
        seen_emails.add(email)
        merged.append({
            "name": name,
            "email": email,
            "title": li.get("title"),
            "company": company or li.get("company"),
            "company_domain": domain,
            "linkedin_url": li.get("linkedin_url"),
            "confidence": "low" if not li.get("email") else "medium",
        })
    return merged


def _parse_csv(content: bytes) -> list[dict]:
    """Parse CSV, auto-detect delimiter. Expects columns: name, email, title, company (email required)."""
    text = content.decode("utf-8-sig", errors="replace")
    reader = csv.DictReader(io.StringIO(text))
    rows = []
    for r in reader:
        row = {k.strip().lower().replace(" ", "_"): v.strip() if v else "" for k, v in r.items()}
        email = row.get("email") or row.get("e-mail") or row.get("email_address")
        if not email or "@" not in email:
            continue
        rows.append({
            "name": row.get("name") or row.get("full_name") or "",
            "email": email,
            "title": row.get("title") or row.get("job_title") or row.get("position") or "",
            "company": row.get("company") or row.get("organization") or "",
            "company_domain": row.get("domain") or row.get("company_domain") or "",
            "linkedin_url": row.get("linkedin") or row.get("linkedin_url") or "",
        })
    return rows


def _parse_excel(content: bytes) -> list[dict]:
    """Parse Excel (.xlsx). Uses first sheet, expects header row with name, email, title, company."""
    try:
        from openpyxl import load_workbook
    except ImportError:
        raise HTTPException(500, "openpyxl not installed")
    wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    ws = wb.active
    if not ws:
        return []
    rows_iter = ws.iter_rows(values_only=True)
    headers = [str(h).strip().lower().replace(" ", "_") if h else "" for h in next(rows_iter, [])]
    col_map = {h: i for i, h in enumerate(headers) if h}
    email_col = col_map.get("email") or col_map.get("e-mail") or col_map.get("email_address")
    if email_col is None:
        email_col = next((i for i, h in enumerate(headers) if h and "email" in h), None)
    if email_col is None:
        return []
    rows = []
    name_col = col_map.get("name") or col_map.get("full_name")
    title_col = col_map.get("title") or col_map.get("job_title") or col_map.get("position")
    company_col = col_map.get("company") or col_map.get("organization")
    domain_col = col_map.get("domain") or col_map.get("company_domain")
    linkedin_col = col_map.get("linkedin") or col_map.get("linkedin_url")
    for row in rows_iter:
        vals = list(row) if row else []
        email = (vals[email_col] if email_col is not None and email_col < len(vals) else "") or ""
        if not email or "@" not in str(email):
            continue
        rows.append({
            "name": str(vals[name_col] or "") if name_col is not None and name_col < len(vals) else "",
            "email": str(email).strip(),
            "title": str(vals[title_col] or "") if title_col is not None and title_col < len(vals) else "",
            "company": str(vals[company_col] or "") if company_col is not None and company_col < len(vals) else "",
            "company_domain": str(vals[domain_col] or "") if domain_col is not None and domain_col < len(vals) else "",
            "linkedin_url": str(vals[linkedin_col] or "") if linkedin_col is not None and linkedin_col < len(vals) else "",
        })
    return rows


@router.post("/import")
async def import_contacts(
    file: UploadFile = File(...),
    skip_duplicates: bool = True,
    user: dict | None = Depends(get_current_user_optional),
):
    """Import contacts from CSV or Excel. Duplicates (by email) are skipped by default; set skip_duplicates=false to get errors on duplicate."""
    content = await file.read()
    filename = (file.filename or "").lower()
    if filename.endswith(".csv"):
        rows = _parse_csv(content)
    elif filename.endswith(".xlsx"):
        rows = _parse_excel(content)
    else:
        raise HTTPException(400, "Upload CSV or Excel (.xlsx)")
    if not rows:
        raise HTTPException(400, "No valid contacts found. Ensure file has 'email' column and at least one row.")
    db = await get_db()
    created = []
    duplicates_skipped = 0
    try:
        for c in rows:
            try:
                email_clean = sanitize_email(c.get("email") or "")
                domain_clean = normalize_domain(c.get("company_domain") or "")
                if skip_duplicates:
                    cursor = await db.execute("SELECT id FROM contacts WHERE email = ?", (email_clean,))
                    if await cursor.fetchone():
                        duplicates_skipped += 1
                        continue
                cursor = await db.execute(
                    """INSERT INTO contacts (name, email, title, company, company_domain, linkedin_url, confidence, department)
                       VALUES (?, ?, ?, ?, ?, ?, 'medium', ?)""",
                    (
                        c.get("name") or "Unknown",
                        email_clean,
                        c.get("title"),
                        c.get("company"),
                        domain_clean,
                        c.get("linkedin_url"),
                        None,
                    ),
                )
                row_id = cursor.lastrowid
                await db.commit()
                created.append({"id": row_id, **c})
            except Exception:
                await db.rollback()
                if not skip_duplicates:
                    raise
                pass
    finally:
        await db.close()
    if user:
        await log_event(
            user["id"], "scrape_completed", "scraper",
            {"source": "import", "count": len(created), "duplicates_skipped": duplicates_skipped},
        )
    return {"contacts": created, "count": len(created), "duplicates_skipped": duplicates_skipped}


@router.post("/search-person")
async def search_person(req: SearchPersonRequest):
    """Search the web for information about a person (name + optional company). Uses Tavily if TAVILY_API_KEY is set; optional LLM summary via Ollama."""
    query = req.name.strip()
    if req.company and req.company.strip():
        query = f"{query} {req.company.strip()}"
    if not query:
        raise HTTPException(400, "Name is required")

    api_key = (os.getenv("TAVILY_API_KEY") or "").strip()
    if not api_key:
        return {
            "query": query,
            "results": [],
            "summary": None,
            "message": "Web search is not configured. Set TAVILY_API_KEY in the backend .env to enable finding contact information from the internet.",
        }

    import httpx
    async with httpx.AsyncClient(timeout=30.0) as client:
        try:
            r = await client.post(
                "https://api.tavily.com/search",
                json={
                    "api_key": api_key,
                    "query": f"{req.name} contact email professional {req.company or ''}".strip(),
                    "search_depth": "basic",
                    "max_results": 10,
                },
            )
            r.raise_for_status()
            data = r.json()
        except httpx.HTTPStatusError as e:
            return {
                "query": query,
                "results": [],
                "summary": None,
                "message": f"Search API error: {e.response.status_code}",
            }
        except Exception as e:
            return {
                "query": query,
                "results": [],
                "summary": None,
                "message": str(e),
            }

    results = [
        {"title": x.get("title"), "url": x.get("url"), "content": (x.get("content") or "")[:500]}
        for x in data.get("results") or []
    ]

    # Optional: LLM summary via Ollama
    summary = None
    ollama_url = os.getenv("OLLAMA_URL", "http://localhost:11434")
    if results:
        snippets = "\n\n".join(
            f"[{i+1}] {r.get('title', '')}\n{r.get('content', '')}" for i, r in enumerate(results[:6])
        )
        prompt = f"""Based on the following web search results about "{req.name}"{f' at {req.company}' if req.company else ''}, extract and list:
- Possible job title and company
- Email or contact info if mentioned
- LinkedIn or social profile URLs if mentioned
- One short paragraph summarizing who they are and relevance for outreach

Search results:
{snippets}

Respond in clear bullet points and one short paragraph. If no contact info is found, say so."""

        try:
            async with httpx.AsyncClient(timeout=60.0) as client_ollama:
                resp = await client_ollama.post(
                    f"{ollama_url.rstrip('/')}/api/generate",
                    json={"model": "llama3.2", "prompt": prompt, "stream": False},
                )
                if resp.status_code == 200:
                    body = resp.json()
                    summary = (body.get("response") or "").strip()
        except Exception:
            pass

    return {
        "query": query,
        "results": results,
        "summary": summary,
        "message": None,
    }


@router.post("/scrape")
async def scrape_contacts(req: ScrapeRequest):
    """Scrape contacts from domain, company name, and/or LinkedIn company URL."""
    domain = req.domain
    if not domain and req.company_name:
        domain = extract_domain_from_company(req.company_name)
    if not domain and not req.linkedin_url:
        raise HTTPException(400, "Provide domain, company_name, or linkedin_url")

    domain_contacts = []
    if domain:
        domain_contacts = await scrape_contacts_from_domain(
            domain=domain,
            company_name=req.company_name,
        )

    linkedin_contacts = []
    company_name = req.company_name
    if req.linkedin_url:
        li_data = await scrape_linkedin_company(
            req.linkedin_url,
            max_employees=req.linkedin_max_employees or 50,
        )
        company_name = company_name or li_data.get("company_name")
        if li_data.get("contacts"):
            linkedin_contacts = li_data["contacts"]

    if not domain and company_name:
        domain = extract_domain_from_company(company_name)
    domain = normalize_domain(domain or "")

    custom_patterns = []
    db_prep = await get_db()
    try:
        cursor = await db_prep.execute("SELECT pattern FROM custom_email_formats ORDER BY priority DESC")
        rows = await cursor.fetchall()
        custom_patterns = [r["pattern"] for r in rows if r.get("pattern")]
    except Exception:
        pass
    finally:
        await db_prep.close()

    contacts_data = _merge_contacts(
        domain_contacts, linkedin_contacts, company_name or "", domain or "", custom_patterns
    )

    db = await get_db()
    created = []
    duplicates_skipped = 0
    try:
        for c in contacts_data:
            try:
                email_clean = sanitize_email(c.get("email") or "")
                domain_clean = normalize_domain(c.get("company_domain") or "")
                cursor = await db.execute("SELECT id FROM contacts WHERE email = ?", (email_clean,))
                if await cursor.fetchone():
                    duplicates_skipped += 1
                    continue
                cursor = await db.execute(
                    """INSERT INTO contacts (name, email, title, company, company_domain, linkedin_url, confidence, department)
                       VALUES (?, ?, ?, ?, ?, ?, ?, ?)""",
                    (
                        c.get("name"),
                        email_clean,
                        c.get("title"),
                        c.get("company"),
                        domain_clean,
                        c.get("linkedin_url"),
                        c.get("confidence", "medium"),
                        c.get("department"),
                    ),
                )
                row_id = cursor.lastrowid
                await db.commit()
                created.append({"id": row_id, **c, "email": email_clean, "company_domain": domain_clean})
            except Exception:
                await db.rollback()
                pass
    finally:
        await db.close()
    return {"contacts": created, "count": len(created), "duplicates_skipped": duplicates_skipped}


@router.get("")
async def list_contacts(
    company: str | None = None,
    q: str | None = None,
    pipeline_status: str | None = None,
    limit: int = 500,
    mine_only: bool = False,
    user: dict | None = Depends(get_current_user_optional),
):
    """List contacts. Optional q (search name/email/company), pipeline_status filter. Standard users see only their contacts + unassigned. Admins see all."""
    db = await get_db()
    try:
        conditions, params = [], []
        if company:
            conditions.append("company LIKE ?")
            params.append(f"%{company}%")
        if q and q.strip():
            q_term = f"%{q.strip()}%"
            conditions.append("(name LIKE ? OR email LIKE ? OR company LIKE ? OR title LIKE ?)")
            params.extend([q_term, q_term, q_term, q_term])
        if pipeline_status and pipeline_status.strip():
            conditions.append("(pipeline_status = ? OR (pipeline_status IS NULL AND ? = 'cold'))")
            params.append(pipeline_status.strip().lower(), pipeline_status.strip().lower())
        if user and user.get("role") != "admin":
            if mine_only:
                conditions.append("owner_id = ?")
                params.append(user["id"])
            else:
                conditions.append("(owner_id = ? OR owner_id IS NULL)")
                params.append(user["id"])
        where = ("WHERE " + " AND ".join(conditions)) if conditions else ""
        params.append(limit)
        cursor = await db.execute(
            f"SELECT * FROM contacts {where} ORDER BY created_at DESC LIMIT ?",
            params,
        )
        rows = await cursor.fetchall()
        result = [dict(r) for r in rows]
        for r in result:
            if r.get("email"):
                r["email"] = sanitize_email(r["email"])
            if r.get("company_domain"):
                r["company_domain"] = normalize_domain(r["company_domain"])
        return result
    finally:
        await db.close()


@router.post("")
async def create_contact(contact: ContactCreate, user: dict | None = Depends(get_current_user_optional)):
    """Manually add a contact."""
    db = await get_db()
    try:
        cursor = await db.execute(
            """INSERT INTO contacts (name, email, title, company, company_domain, linkedin_url, confidence, department)
               VALUES (?, ?, ?, ?, ?, ?, ?, ?)""",
            (
                contact.name,
                contact.email,
                contact.title,
                contact.company,
                contact.company_domain,
                contact.linkedin_url,
                contact.confidence or "medium",
                contact.department,
            ),
        )
        await db.commit()
        row_id = cursor.lastrowid
        cursor = await db.execute("SELECT * FROM contacts WHERE id = ?", (row_id,))
        row = await cursor.fetchone()
        d = dict(row)
        if d.get("email"):
            d["email"] = sanitize_email(d["email"])
        if user:
            await log_audit(user["id"], "contact_create", "contact", str(row_id), contact.email)
        return d
    except Exception as e:
        await db.rollback()
        raise HTTPException(400, str(e))
    finally:
        await db.close()


@router.get("/{contact_id}")
async def get_contact(contact_id: int):
    """Get a single contact."""
    db = await get_db()
    try:
        cursor = await db.execute("SELECT * FROM contacts WHERE id = ?", (contact_id,))
        row = await cursor.fetchone()
        if not row:
            raise HTTPException(404, "Contact not found")
        d = dict(row)
        if d.get("email"):
            d["email"] = sanitize_email(d["email"])
        if d.get("company_domain"):
            d["company_domain"] = normalize_domain(d["company_domain"])
        return d
    finally:
        await db.close()


@router.post("/fix-emails")
async def fix_malformed_emails():
    """Fix contacts with malformed emails (e.g. name@https://domain.com/path -> name@domain.com)."""
    db = await get_db()
    try:
        cursor = await db.execute("SELECT id, email, company_domain FROM contacts")
        rows = await cursor.fetchall()
        fixed = 0
        for r in rows:
            email_new = sanitize_email(r["email"] or "")
            domain_new = normalize_domain(r["company_domain"] or "")
            if email_new != (r["email"] or "") or domain_new != (r["company_domain"] or ""):
                await db.execute(
                    "UPDATE contacts SET email = ?, company_domain = ? WHERE id = ?",
                    (email_new, domain_new, r["id"]),
                )
                fixed += 1
        await db.commit()
        return {"fixed": fixed, "message": f"Updated {fixed} contacts"}
    finally:
        await db.close()


@router.delete("/{contact_id}")
async def delete_contact(contact_id: int, user: dict | None = Depends(get_current_user_optional)):
    """Delete a contact."""
    db = await get_db()
    try:
        cursor = await db.execute("SELECT email FROM contacts WHERE id = ?", (contact_id,))
        row = await cursor.fetchone()
        await db.execute("DELETE FROM contacts WHERE id = ?", (contact_id,))
        await db.commit()
        if user and row:
            await log_audit(user["id"], "contact_delete", "contact", str(contact_id), row.get("email", ""))
        return {"ok": True}
    finally:
        await db.close()