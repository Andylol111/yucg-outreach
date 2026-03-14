"""
Admin API - User management, audit log, API keys, notification prefs, 2FA.
Admin-only access.
"""
import io
import secrets
import hashlib
from fastapi import APIRouter, HTTPException, Depends
from fastapi.responses import Response
from pydantic import BaseModel
from typing import Optional

from app.database import get_db, row_to_dict
from app.auth_deps import get_current_admin
from app.services.audit_service import log_audit

# YUCG palette for Excel styling
YUCG_DEEP_NAVY = "1A2F5A"
YUCG_MID_NAVY = "1E3A6E"
YUCG_SLATE_BLUE = "3D5C82"
YUCG_STEEL_BLUE = "5B7FA6"
YUCG_PALE_SKY = "C8DCED"
YUCG_WHITE = "FFFFFF"

router = APIRouter()


# --- Login log (admin only) ---
@router.get("/login-log")
async def get_login_log(limit: int = 50, _admin: dict = Depends(get_current_admin)):
    """List recent logins. Admin only."""
    db = await get_db()
    try:
        cursor = await db.execute(
            """SELECT ll.id, ll.email, ll.name, ll.created_at, u.id as user_id
               FROM login_log ll JOIN users u ON ll.user_id = u.id
               ORDER BY ll.created_at DESC LIMIT ?""",
            (limit,),
        )
        rows = await cursor.fetchall()
        return [dict(r) for r in rows]
    finally:
        await db.close()


# --- User management ---
class UserRoleUpdate(BaseModel):
    role: str  # admin | standard


class UserStatusUpdate(BaseModel):
    is_active: bool


class UserInvite(BaseModel):
    email: str


@router.get("/users")
async def list_users(admin: dict = Depends(get_current_admin)):
    """List all users. Admin only."""
    db = await get_db()
    try:
        cursor = await db.execute(
            "SELECT id, email, name, picture, role, is_active, created_at FROM users ORDER BY created_at"
        )
        rows = await cursor.fetchall()
        return [dict(r) for r in rows]
    finally:
        await db.close()


@router.get("/users/export", response_class=Response)
async def export_users_excel(admin: dict = Depends(get_current_admin)):
    """Export all users and profile data to a styled Excel file. Admin only."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, Fill, PatternFill, Alignment, Border, Side

    db = await get_db()
    try:
        cursor = await db.execute(
            """SELECT u.id, u.email, u.name, u.picture, u.role, u.is_active, u.created_at,
                      p.projects, p.experience, p.role_title, p.linkedin_url, p.slack_handle, p.other_handles, p.updated_at AS profile_updated_at
               FROM users u
               LEFT JOIN user_profiles p ON p.user_id = u.id
               ORDER BY u.created_at"""
        )
        rows = await cursor.fetchall()
    finally:
        await db.close()

    wb = Workbook()
    ws = wb.active
    ws.title = "Users"

    headers = [
        "ID", "Email", "Name", "Picture URL", "Role", "Status", "Created At",
        "Projects", "Experience", "Role Title", "LinkedIn URL", "Slack Handle", "Other Handles", "Profile Updated"
    ]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True, color=YUCG_WHITE, size=11)
        cell.fill = PatternFill(start_color=YUCG_DEEP_NAVY, end_color=YUCG_DEEP_NAVY, fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = Border(
            left=Side(style="thin", color=YUCG_SLATE_BLUE),
            right=Side(style="thin", color=YUCG_SLATE_BLUE),
            top=Side(style="thin", color=YUCG_SLATE_BLUE),
            bottom=Side(style="thin", color=YUCG_SLATE_BLUE),
        )

    thin_border = Border(
        left=Side(style="thin", color=YUCG_PALE_SKY),
        right=Side(style="thin", color=YUCG_PALE_SKY),
        top=Side(style="thin", color=YUCG_PALE_SKY),
        bottom=Side(style="thin", color=YUCG_PALE_SKY),
    )
    pale_fill = PatternFill(start_color="E8EEF4", end_color="E8EEF4", fill_type="solid")  # bg-primary tint
    white_fill = PatternFill(start_color=YUCG_WHITE, end_color=YUCG_WHITE, fill_type="solid")

    for r_idx, row in enumerate(rows, 2):
        d = dict(row)
        values = [
            d.get("id"),
            d.get("email") or "",
            d.get("name") or "",
            d.get("picture") or "",
            d.get("role") or "standard",
            "Active" if d.get("is_active", 1) else "Inactive",
            str(d.get("created_at") or ""),
            d.get("projects") or "",
            d.get("experience") or "",
            d.get("role_title") or "",
            d.get("linkedin_url") or "",
            d.get("slack_handle") or "",
            d.get("other_handles") or "",
            str(d.get("profile_updated_at") or ""),
        ]
        for col, val in enumerate(values, 1):
            cell = ws.cell(row=r_idx, column=col, value=val)
            cell.border = thin_border
            cell.fill = pale_fill if r_idx % 2 == 0 else white_fill
            cell.alignment = Alignment(vertical="center", wrap_text=True)
            cell.font = Font(color=YUCG_DEEP_NAVY, size=10)

    from openpyxl.utils import get_column_letter
    for col in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 18
    ws.column_dimensions["B"].width = 28
    ws.column_dimensions["G"].width = 22
    ws.column_dimensions["H"].width = 24
    ws.column_dimensions["N"].width = 22

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return Response(
        content=buf.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=YUCG_users_export.xlsx"},
    )


@router.post("/users/invite")
async def invite_user(payload: UserInvite, admin: dict = Depends(get_current_admin)):
    """Create a placeholder user by email. They get role=standard when they first log in. Admin only."""
    email = payload.email.strip().lower()
    if not email or "@" not in email:
        raise HTTPException(400, "Invalid email")
    if not email.endswith("@yale.edu"):
        raise HTTPException(400, "Only @yale.edu emails allowed")
    db = await get_db()
    try:
        cursor = await db.execute("SELECT id FROM users WHERE email = ?", (email,))
        if await cursor.fetchone():
            raise HTTPException(400, "User already exists")
        await db.execute(
            "INSERT INTO users (email, name, role) VALUES (?, '', 'standard')",
            (email,),
        )
        await db.commit()
        await log_audit(admin["id"], "user_invite", "user", email, f"Invited {email}")
        return {"ok": True, "email": email}
    finally:
        await db.close()


@router.patch("/users/{user_id}/role")
async def update_user_role(user_id: int, payload: UserRoleUpdate, admin: dict = Depends(get_current_admin)):
    """Update user role. Admin only."""
    if payload.role not in ("admin", "standard"):
        raise HTTPException(400, "Role must be admin or standard")
    if admin["id"] == user_id and payload.role != "admin":
        raise HTTPException(400, "Cannot demote yourself")
    db = await get_db()
    try:
        cursor = await db.execute("SELECT email FROM users WHERE id = ?", (user_id,))
        row = await cursor.fetchone()
        if not row:
            raise HTTPException(404, "User not found")
        await db.execute("UPDATE users SET role = ? WHERE id = ?", (payload.role, user_id))
        await db.commit()
        await log_audit(admin["id"], "role_update", "user", str(user_id), f"Set {row['email']} to {payload.role}")
        return {"ok": True, "role": payload.role}
    finally:
        await db.close()


@router.patch("/users/{user_id}/status")
async def update_user_status(user_id: int, payload: UserStatusUpdate, admin: dict = Depends(get_current_admin)):
    """Activate or deactivate user. Admin only."""
    if admin["id"] == user_id and not payload.is_active:
        raise HTTPException(400, "Cannot deactivate yourself")
    db = await get_db()
    try:
        cursor = await db.execute("SELECT email FROM users WHERE id = ?", (user_id,))
        row = await cursor.fetchone()
        if not row:
            raise HTTPException(404, "User not found")
        await db.execute("UPDATE users SET is_active = ? WHERE id = ?", (1 if payload.is_active else 0, user_id))
        await db.commit()
        await log_audit(admin["id"], "status_update", "user", str(user_id), f"{'Activated' if payload.is_active else 'Deactivated'} {row['email']}")
        return {"ok": True, "is_active": payload.is_active}
    finally:
        await db.close()


# --- Audit log ---
@router.get("/audit-log")
async def get_audit_log(limit: int = 100, admin: dict = Depends(get_current_admin)):
    """List audit log entries. Admin only."""
    db = await get_db()
    try:
        cursor = await db.execute(
            """SELECT al.*, u.email as user_email FROM audit_log al
               LEFT JOIN users u ON al.user_id = u.id
               ORDER BY al.created_at DESC LIMIT ?""",
            (limit,),
        )
        rows = await cursor.fetchall()
        return [dict(r) for r in rows]
    finally:
        await db.close()


@router.get("/audit-log/export", response_class=Response)
async def export_audit_log_excel(admin: dict = Depends(get_current_admin)):
    """Export audit log to Excel. Admin only."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    db = await get_db()
    try:
        cursor = await db.execute(
            """SELECT al.id, al.user_id, u.email as user_email, al.action, al.resource_type, al.resource_id, al.details, al.created_at
               FROM audit_log al LEFT JOIN users u ON al.user_id = u.id
               ORDER BY al.created_at DESC LIMIT 5000"""
        )
        rows = await cursor.fetchall()
    finally:
        await db.close()

    wb = Workbook()
    ws = wb.active
    ws.title = "Audit log"
    headers = ["ID", "User ID", "User Email", "Action", "Resource Type", "Resource ID", "Details", "Created At"]
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=col, value=h)
        c.font = Font(bold=True, color=YUCG_WHITE)
        c.fill = PatternFill(start_color=YUCG_DEEP_NAVY, end_color=YUCG_DEEP_NAVY, fill_type="solid")
    for r_idx, row in enumerate(rows, 2):
        d = dict(row)
        for col, key in enumerate(["id", "user_id", "user_email", "action", "resource_type", "resource_id", "details", "created_at"], 1):
            val = d.get(key)
            ws.cell(row=r_idx, column=col, value=str(val) if val is not None else "")
    for col in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 16
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return Response(
        content=buf.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=YUCG_audit_log.xlsx"},
    )


@router.get("/export/all", response_class=Response)
async def export_all_admin(admin: dict = Depends(get_current_admin)):
    """Export all admin data in one ZIP: users, audit log, and operations full export. Admin only."""
    import zipfile

    zip_buf = io.BytesIO()
    with zipfile.ZipFile(zip_buf, "w", zipfile.ZIP_DEFLATED) as zf:
        # 1) Users Excel
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
        db = await get_db()
        try:
            cursor = await db.execute(
                """SELECT u.id, u.email, u.name, u.picture, u.role, u.is_active, u.created_at,
                          p.projects, p.experience, p.role_title, p.linkedin_url, p.slack_handle, p.other_handles, p.updated_at AS profile_updated_at
                   FROM users u LEFT JOIN user_profiles p ON p.user_id = u.id ORDER BY u.created_at"""
            )
            rows = await cursor.fetchall()
        finally:
            await db.close()
        wb = Workbook()
        ws = wb.active
        ws.title = "Users"
        headers = ["ID", "Email", "Name", "Picture URL", "Role", "Status", "Created At", "Projects", "Experience", "Role Title", "LinkedIn URL", "Slack Handle", "Other Handles", "Profile Updated"]
        for col, h in enumerate(headers, 1):
            c = ws.cell(row=1, column=col, value=h)
            c.font = Font(bold=True, color=YUCG_WHITE, size=11)
            c.fill = PatternFill(start_color=YUCG_DEEP_NAVY, end_color=YUCG_DEEP_NAVY, fill_type="solid")
        thin_border = Border(left=Side(style="thin", color=YUCG_PALE_SKY), right=Side(style="thin", color=YUCG_PALE_SKY), top=Side(style="thin", color=YUCG_PALE_SKY), bottom=Side(style="thin", color=YUCG_PALE_SKY))
        pale_fill = PatternFill(start_color="E8EEF4", end_color="E8EEF4", fill_type="solid")
        white_fill = PatternFill(start_color=YUCG_WHITE, end_color=YUCG_WHITE, fill_type="solid")
        for r_idx, row in enumerate(rows, 2):
            d = dict(row)
            values = [d.get("id"), d.get("email") or "", d.get("name") or "", d.get("picture") or "", d.get("role") or "standard", "Active" if d.get("is_active", 1) else "Inactive", str(d.get("created_at") or ""), d.get("projects") or "", d.get("experience") or "", d.get("role_title") or "", d.get("linkedin_url") or "", d.get("slack_handle") or "", d.get("other_handles") or "", str(d.get("profile_updated_at") or "")]
            for col, val in enumerate(values, 1):
                cell = ws.cell(row=r_idx, column=col, value=val)
                cell.border = thin_border
                cell.fill = pale_fill if r_idx % 2 == 0 else white_fill
                cell.font = Font(color=YUCG_DEEP_NAVY, size=10)
        for col in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(col)].width = 18
        u_buf = io.BytesIO()
        wb.save(u_buf)
        u_buf.seek(0)
        zf.writestr("users.xlsx", u_buf.getvalue())

        # 2) Audit log Excel
        db = await get_db()
        try:
            cursor = await db.execute(
                """SELECT al.id, al.user_id, u.email as user_email, al.action, al.resource_type, al.resource_id, al.details, al.created_at
                   FROM audit_log al LEFT JOIN users u ON al.user_id = u.id ORDER BY al.created_at DESC LIMIT 5000"""
            )
            audit_rows = await cursor.fetchall()
        finally:
            await db.close()
        wb2 = Workbook()
        ws2 = wb2.active
        ws2.title = "Audit log"
        audit_headers = ["ID", "User ID", "User Email", "Action", "Resource Type", "Resource ID", "Details", "Created At"]
        for col, h in enumerate(audit_headers, 1):
            c = ws2.cell(row=1, column=col, value=h)
            c.font = Font(bold=True, color=YUCG_WHITE)
            c.fill = PatternFill(start_color=YUCG_DEEP_NAVY, end_color=YUCG_DEEP_NAVY, fill_type="solid")
        for r_idx, row in enumerate(audit_rows, 2):
            d = dict(row)
            for col, key in enumerate(["id", "user_id", "user_email", "action", "resource_type", "resource_id", "details", "created_at"], 1):
                val = d.get(key)
                ws2.cell(row=r_idx, column=col, value=str(val) if val is not None else "")
        for col in range(1, len(audit_headers) + 1):
            ws2.column_dimensions[get_column_letter(col)].width = 16
        a_buf = io.BytesIO()
        wb2.save(a_buf)
        a_buf.seek(0)
        zf.writestr("audit_log.xlsx", a_buf.getvalue())

        # 3) Operations full export as a nested ZIP (so one download has everything)
        from app.routers.operations import build_operations_full_export_bytes
        try:
            ops_bytes = await build_operations_full_export_bytes(admin, 30)
            zf.writestr("operations_full_export_30d.zip", ops_bytes)
        except Exception:
            pass  # if operations export fails, still return users + audit

    zip_buf.seek(0)
    return Response(
        content=zip_buf.getvalue(),
        media_type="application/zip",
        headers={"Content-Disposition": "attachment; filename=YUCG_admin_export_all.zip"},
    )


# --- API keys ---
class ApiKeyCreate(BaseModel):
    name: str
    scopes: Optional[str] = None  # e.g. "contacts:read"


@router.get("/api-keys")
async def list_api_keys(admin: dict = Depends(get_current_admin)):
    """List API keys for current user. Admin only."""
    db = await get_db()
    try:
        cursor = await db.execute(
            "SELECT id, key_prefix, name, scopes, created_at, last_used_at FROM api_keys WHERE user_id = ?",
            (admin["id"],),
        )
        rows = await cursor.fetchall()
        return [dict(r) for r in rows]
    finally:
        await db.close()


@router.post("/api-keys")
async def create_api_key(payload: ApiKeyCreate, admin: dict = Depends(get_current_admin)):
    """Create a new API key. Returns the raw key once (store it securely). Admin only."""
    raw_key = "yucg_" + secrets.token_urlsafe(32)
    key_hash = hashlib.sha256(raw_key.encode()).hexdigest()
    key_prefix = raw_key[:16] + "..."
    db = await get_db()
    try:
        await db.execute(
            "INSERT INTO api_keys (user_id, key_hash, key_prefix, name, scopes) VALUES (?, ?, ?, ?, ?)",
            (admin["id"], key_hash, key_prefix, payload.name, payload.scopes or "contacts:read"),
        )
        await db.commit()
        await log_audit(admin["id"], "api_key_create", "api_key", payload.name, "Created new API key")
        return {"key": raw_key, "key_prefix": key_prefix, "name": payload.name, "warning": "Store this key securely. It will not be shown again."}
    finally:
        await db.close()


@router.delete("/api-keys/{key_id}")
async def revoke_api_key(key_id: int, admin: dict = Depends(get_current_admin)):
    """Revoke an API key. Admin only."""
    db = await get_db()
    try:
        cursor = await db.execute(
            "DELETE FROM api_keys WHERE id = ? AND user_id = ?",
            (key_id, admin["id"]),
        )
        await db.commit()
        if cursor.rowcount == 0:
            raise HTTPException(404, "API key not found")
        await log_audit(admin["id"], "api_key_revoke", "api_key", str(key_id), "Revoked API key")
        return {"ok": True}
    finally:
        await db.close()


# --- Projects & team assignments (semester + client, e.g. Spring 2026 - Project Lego) ---
class ProjectCreate(BaseModel):
    name: str
    semester: Optional[str] = None
    description: Optional[str] = None


class UserProjectAssign(BaseModel):
    project_id: int
    role_in_project: Optional[str] = None


@router.get("/projects")
async def list_projects(admin: dict = Depends(get_current_admin)):
    """List all projects. Admin only."""
    db = await get_db()
    try:
        cursor = await db.execute(
            "SELECT id, name, semester, description, created_at FROM projects ORDER BY semester DESC, name"
        )
        rows = await cursor.fetchall()
        return [row_to_dict(r) for r in rows]
    finally:
        await db.close()


@router.post("/projects")
async def create_project(payload: ProjectCreate, admin: dict = Depends(get_current_admin)):
    """Create a project (e.g. Spring 2026 - Project Lego). Admin only."""
    name = payload.name.strip()
    if not name:
        raise HTTPException(400, "Project name required")
    db = await get_db()
    try:
        cursor = await db.execute(
            "INSERT INTO projects (name, semester, description) VALUES (?, ?, ?)",
            (name, (payload.semester or "").strip() or None, (payload.description or "").strip() or None),
        )
        await db.commit()
        pid = cursor.lastrowid
        await log_audit(admin["id"], "project_create", "project", str(pid), f"Created project: {name}")
        return {"id": pid, "name": name, "semester": payload.semester, "description": payload.description}
    finally:
        await db.close()


@router.delete("/projects/{project_id}")
async def delete_project(project_id: int, admin: dict = Depends(get_current_admin)):
    """Delete a project. Admin only."""
    db = await get_db()
    try:
        await db.execute("DELETE FROM user_project_assignments WHERE project_id = ?", (project_id,))
        await db.execute("DELETE FROM projects WHERE id = ?", (project_id,))
        await db.commit()
        await log_audit(admin["id"], "project_delete", "project", str(project_id), "Deleted project")
        return {"ok": True}
    finally:
        await db.close()


@router.get("/projects/{project_id}/assignments")
async def list_project_assignments(project_id: int, admin: dict = Depends(get_current_admin)):
    """List users assigned to a project. Admin only."""
    db = await get_db()
    try:
        cursor = await db.execute(
            """SELECT u.id, u.email, u.name, upa.role_in_project
               FROM user_project_assignments upa
               JOIN users u ON u.id = upa.user_id
               WHERE upa.project_id = ?
               ORDER BY u.name""",
            (project_id,),
        )
        rows = await cursor.fetchall()
        return [row_to_dict(r) for r in rows]
    finally:
        await db.close()


@router.put("/users/{user_id}/project")
async def assign_user_to_project(user_id: int, payload: UserProjectAssign, admin: dict = Depends(get_current_admin)):
    """Assign user to project with role. Admin only."""
    db = await get_db()
    try:
        await db.execute(
            """INSERT OR REPLACE INTO user_project_assignments (user_id, project_id, role_in_project)
               VALUES (?, ?, ?)""",
            (user_id, payload.project_id, (payload.role_in_project or "").strip() or None),
        )
        await db.commit()
        await log_audit(admin["id"], "project_assign", "user", str(user_id), f"Assigned to project {payload.project_id}")
        return {"ok": True}
    finally:
        await db.close()


@router.delete("/users/{user_id}/project/{project_id}")
async def unassign_user_from_project(user_id: int, project_id: int, admin: dict = Depends(get_current_admin)):
    """Remove user from project. Admin only."""
    db = await get_db()
    try:
        await db.execute(
            "DELETE FROM user_project_assignments WHERE user_id = ? AND project_id = ?",
            (user_id, project_id),
        )
        await db.commit()
        await log_audit(admin["id"], "project_unassign", "user", str(user_id), f"Removed from project {project_id}")
        return {"ok": True}
    finally:
        await db.close()


@router.get("/users/{user_id}/projects")
async def list_user_projects(user_id: int, admin: dict = Depends(get_current_admin)):
    """List projects assigned to a user. Admin only."""
    db = await get_db()
    try:
        cursor = await db.execute(
            """SELECT p.id, p.name, p.semester, upa.role_in_project
               FROM user_project_assignments upa
               JOIN projects p ON p.id = upa.project_id
               WHERE upa.user_id = ?
               ORDER BY p.semester DESC, p.name""",
            (user_id,),
        )
        rows = await cursor.fetchall()
        return [row_to_dict(r) for r in rows]
    finally:
        await db.close()


# --- 2FA for admins ---
@router.get("/2fa/status")
async def get_2fa_status(admin: dict = Depends(get_current_admin)):
    """Get 2FA status: enabled, pending (setup in progress), or not_setup. Admin only."""
    db = await get_db()
    try:
        cursor = await db.execute("SELECT totp_secret FROM users WHERE id = ?", (admin["id"],))
        row = await cursor.fetchone()
        if row and row["totp_secret"]:
            return {"status": "enabled"}
        cursor = await db.execute("SELECT 1 FROM pending_2fa_setup WHERE user_id = ?", (admin["id"],))
        if await cursor.fetchone():
            return {"status": "pending"}
        return {"status": "not_setup"}
    finally:
        await db.close()


@router.post("/2fa/setup")
async def setup_2fa(admin: dict = Depends(get_current_admin)):
    """Generate TOTP secret and return QR/provisioning URI. Secret stored in pending until verified. Admin only."""
    try:
        import pyotp
    except ImportError:
        raise HTTPException(500, "pyotp not installed")
    db = await get_db()
    try:
        cursor = await db.execute("SELECT totp_secret FROM users WHERE id = ?", (admin["id"],))
        row = await cursor.fetchone()
        if row and row["totp_secret"]:
            raise HTTPException(400, "2FA already enabled. Disable first to re-setup.")
        secret = pyotp.random_base32()
        totp = pyotp.TOTP(secret)
        provisioning_uri = totp.provisioning_uri(
            name=admin.get("email", "admin"),
            issuer_name="YUCG Outreach",
        )
        await db.execute(
            "INSERT OR REPLACE INTO pending_2fa_setup (user_id, secret, created_at) VALUES (?, ?, CURRENT_TIMESTAMP)",
            (admin["id"], secret),
        )
        await db.commit()
        await log_audit(admin["id"], "2fa_setup", "user", str(admin["id"]), "Initiated 2FA setup")
        return {"secret": secret, "provisioning_uri": provisioning_uri, "message": "Scan with authenticator app, then verify with /2fa/verify"}
    finally:
        await db.close()


class TwoFactorVerify(BaseModel):
    code: str


@router.post("/2fa/verify")
async def verify_2fa(payload: TwoFactorVerify, admin: dict = Depends(get_current_admin)):
    """Verify TOTP code to complete 2FA setup. Moves secret from pending to users. Admin only."""
    try:
        import pyotp
    except ImportError:
        raise HTTPException(500, "pyotp not installed")
    db = await get_db()
    try:
        cursor = await db.execute("SELECT secret FROM pending_2fa_setup WHERE user_id = ?", (admin["id"],))
        row = await cursor.fetchone()
        if not row:
            raise HTTPException(400, "2FA not set up. Click Setup 2FA first and scan the QR code.")
        secret = row["secret"]
        totp = pyotp.TOTP(secret)
        if not totp.verify(payload.code, valid_window=1):
            raise HTTPException(400, "Invalid code")
        await db.execute("UPDATE users SET totp_secret = ? WHERE id = ?", (secret, admin["id"]))
        await db.execute("DELETE FROM pending_2fa_setup WHERE user_id = ?", (admin["id"],))
        await db.commit()
        await log_audit(admin["id"], "2fa_verified", "user", str(admin["id"]), "2FA enabled")
        return {"ok": True, "message": "2FA enabled"}
    finally:
        await db.close()


@router.post("/2fa/disable")
async def disable_2fa(payload: TwoFactorVerify, admin: dict = Depends(get_current_admin)):
    """Disable 2FA. Requires current TOTP code. Admin only."""
    try:
        import pyotp
    except ImportError:
        raise HTTPException(500, "pyotp not installed")
    db = await get_db()
    try:
        cursor = await db.execute("SELECT totp_secret FROM users WHERE id = ?", (admin["id"],))
        row = await cursor.fetchone()
        if not row or not row["totp_secret"]:
            raise HTTPException(400, "2FA not enabled")
        totp = pyotp.TOTP(row["totp_secret"])
        if not totp.verify(payload.code, valid_window=1):
            raise HTTPException(400, "Invalid code")
        await db.execute("UPDATE users SET totp_secret = NULL WHERE id = ?", (admin["id"],))
        await db.execute("DELETE FROM pending_2fa_setup WHERE user_id = ?", (admin["id"],))
        await db.commit()
        await log_audit(admin["id"], "2fa_disable", "user", str(admin["id"]), "2FA disabled")
        return {"ok": True}
    finally:
        await db.close()


@router.post("/2fa/reset")
async def reset_2fa(admin: dict = Depends(get_current_admin)):
    """Clear 2FA and any pending setup. Use if stuck (e.g. lost phone before verifying). Admin only."""
    db = await get_db()
    try:
        await db.execute("UPDATE users SET totp_secret = NULL WHERE id = ?", (admin["id"],))
        await db.execute("DELETE FROM pending_2fa_setup WHERE user_id = ?", (admin["id"],))
        await db.commit()
        await log_audit(admin["id"], "2fa_reset", "user", str(admin["id"]), "2FA reset (cleared)")
        return {"ok": True, "message": "2FA reset. You can set up again."}
    finally:
        await db.close()
